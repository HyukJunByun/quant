import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import xlwings as xw
import time
import gc

kisrating_url = 'https://www.kisrating.com/ratingsStatistics/statics_spread.do'
webbpage = requests.get(kisrating_url)
webb_data = BeautifulSoup(webbpage.content, 'lxml')
# 회사채 수익률 웹페이지 가져오기
bbb = webb_data.find('div', {'class': 'table_ty1'})
# 회사채 수익률 표 가져오기
webb_data.decompose()
# 정보 빼고 나면 웹페이지는 삭제(스피드업)
webbpage.close()
webbpage = None
bbb_table = bbb.find_all('td')
# 회사채 수익률 표 숫자만 전부 가져오기
bbb_data = bbb_table[98].text
# 회사채 BBB- 5년 수익률 가져오기 = 요구수익률


code_data = pd.read_html('http://kind.krx.co.kr/corpgeneral/corpList.do?method=download&searchType=13', header=0)[0]
start = time.time()
# 시작 시간

def make_code(x):
    x = str(x)
    return '0' * (6 - len(x)) + x


def hms(s):
    hours = s // 3600
    s = s - hours*3600
    mu = s // 60
    ss = s - mu*60
    print('소요시간 = ', hours, '시간 ', mu, '분 ', ss, '초')


buy_zoo = []
# 매수할 주식 목록

code_data = code_data['종목코드']
# ['종목코드'] 요것만 하면 판다스에 종목코드가 리스트로 저장될까? ㅇㅇ 가능하네
code_data = code_data.apply(make_code)

# len(code_data)
for a in range(0, len(code_data)):
    # print(a)
    # 에러확인용
    code_num = code_data[a]
    fnguide_url = "http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A" + code_num + "&cID=&MenuYn=Y&ReportGB=D&NewMenuID=Y&stkGb=701"
    # 종목코드 표에 있는 모든 법인 접속해봄, 일단 100종목

# fnguide_url = "https://comp.fnguide.com/SVO2/ASP/SVD_main.asp?pGB=1&gicode=A005930&cID=&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=&strResearchYN="
# 테스트 쉬우라고 일단 삼성전자 url

    webpage = requests.get(fnguide_url)
    web_data = BeautifulSoup(webpage.content, 'lxml')
    # 웹페이지 정보 가져오기
    try:
        corp_name = web_data.find('h1', {'id': 'giName'}).text
        # 주식이름            +++태그 하나만 가져와도 {}를 붙여야하네, 참고로 이거는 태그까지 전부 포함한 데이터라 .text로 문자만 뽑아내야함
        price_and_num = web_data.find('div', {'class': 'um_table', 'id': 'svdMainGrid1'})
        # 시세현황표
        price_and_num_data = price_and_num.find_all('td')
        # 시세현황표 안에 있는 숫자들 전부 가져오기
        """
        for i in range(0, len(price_and_num_data)):
            print(price_and_num_data[i].text)
        #시세현황표에 있는 숫자들을 find_all 하면 표의 왼쪽에서 오른쪽으로, 위에서 아래순으로 저장한다.
        """
        my_zoo = web_data.find('div', {'class': 'um_table', 'id': 'svdMainGrid5'})
        # 주주구분 현황
        my_zoo_data = my_zoo.find_all('td')
        # 주주구분 현황 안에 있는 숫자들 전부 가져오기
        ifrs_D_A = web_data.find('div', {'class': 'um_table', 'id': 'highlight_D_Y'})
        # ifrs(연결-연간) 데이터 불러오기
        ifrs_D_Q = web_data.find('div', {'class': 'um_table', 'id': 'highlight_D_Q'})
        # ifrs(연결-분기) 데이터 불러오기
        ifrs_B_Q = web_data.find('div', {'class': 'um_table', 'id': 'highlight_B_Q'})
        # ifrs(개별-분기) 데이터 불러오기
        web_data.decompose()
        # 분석 끝낸 웹페이지는 삭제
        webpage.close()
        webpage = None
        #웹페이지 삭제
        gc.collect()
        #불필요한 데이터 전부 삭제
        ifrs_DA = ifrs_D_A.find_all('td')
        ifrs_DQ = ifrs_D_Q.find_all('td')
        ifrs_BQ = ifrs_B_Q.find_all('td')
        # ifrs 숫자들 전부 가져오기
        ifrs_DA_date = ifrs_D_A.find_all('th', {'scope': 'col'})
        # ifrs(연결-연간) 날짜 데이터
        ifrs_DQ_date = ifrs_D_Q.find_all('th', {'scope': 'col'})
        # ifrs(연결-분기) 날짜 데이터
        ifrs_BQ_date = ifrs_B_Q.find_all('th', {'scope': 'col'})

        # ifrs(개별-분기) 날짜 데이터
        if(ifrs_DA[138].text != ifrs_DA[199].text):
            # (마지막 사업보고서 기준) 2년전 roe 정보 없으면 취급하지 않음
            # 199는 연결-연간 표에서 가장 먼 컨센서스 배당수익률. 항상 빈칸이리 카더라
            if(web_data.find_all('span', {'class': "stxt stxt2"})[0].text != 'FICS  창업투자 및 종금'):
                # 기업인수목적회사 정보가 너무 부족해서 걍 거르기
                wb = load_workbook('G:\Hyuk_Rim_v5.xlsx')                             # 엑셀 파일 이름 바뀔때마다 업데이트!!
                # 계산할 엑셀 파일 불러오기
                wb_result = wb['Result']
                # result 워크시트
                wb_data = wb['Data']
                # Data 워크시트

                wb_result['D11'] = bbb_data
                # 요구수익률 넣기
                wb_data['B4'] = corp_name
                # 주식이름 넣기
                wb_data['i5'] = price_and_num_data[0].text
                # 주식 현재가 넣기
                wb_data['i7'] = price_and_num_data[10].text
                # 발행주식수 넣기
                wb_data['i8'] = my_zoo_data[17].text
                # 자기주식수 넣기

                line = 0 # ifrs 표 안에 데이터 순번. 표 안의 모든 데이터를 추출하기 위함
                row = 27 # 엑셀 ifrs(연결-연간)표 첫번째 셀의 행 번호
                for b in range(0, 12):
                    # range 12는 ifrs 표의 행(매출액~자본금) 개수(세로 길이)
                    samsung = 66                                                      # ifrs(연결-연간) 숫자 데이터(매출액~자본금) 엑셀에 넣기
                    # samsung -> chr(66)이 대문자 알파벳 B를 뜻함, ifrs 표의 첫번째 셀 열 번호
                    for c in range(0, 8):
                        # range 8은 ifrs 표의 열 개수(가로 길이)
                        rrow = str(row)
                        col = chr(samsung)
                        cell = col + rrow
                        wb_data[cell] = ifrs_DA[line].text
                        line += 1
                        samsung += 1
                    row += 1

                line = line + 4 * 8
                # ifrs 표에서 중간 내용 건너뛰고 roa 정보 순번
                for bbbb in range(0, 5):
                    samsung = 66                                                      # ifrs(연결-연간) 숫자 데이터(ROA~DPS) 엑셀에 넣기
                    for c in range(0, 8):
                        rrow = str(row)
                        col = chr(samsung)
                        cell = col + rrow
                        wb_data[cell] = ifrs_DA[line].text
                        line += 1
                        samsung += 1
                    row += 1

                line = line + 3 * 8
                samsung = 66
                for cc in range(0, 8):
                        rrow = str(row)                                               # frs(연결-연간) 숫자 데이터(배당수익률) 엑셀에 넣기
                        col = chr(samsung)
                        cell = col + rrow
                        wb_data[cell] = ifrs_DA[line].text
                        line += 1
                        samsung += 1


                for d in range(0, 5):                                                 # ifrs(연결-연간) 날짜 데이터 엑셀에 넣기
                    # range 5 -> 날짜 넣는칸의 총 개수
                    e = d + 66
                    # 날짜 넣는칸의 열 위치(알파벳)
                    g = chr(e) + str(25)
                    # str(25) -> 날짜 넣는칸의 행 위치
                    wb_data[g] = ifrs_DA_date[d + 2].text

                line = 0
                row = 51
                for b in range(0, 12):                                                # ifrs(연결-분기) 숫자 데이터 엑셀에 넣기
                    samsung = 66
                    for c in range(0, 8):
                        rrow = str(row)
                        col = chr(samsung)
                        cell = col + rrow
                        wb_data[cell] = ifrs_DQ[line].text
                        line += 1
                        samsung += 1
                    row += 1


                for d in range(0, 5):                                                 # ifrs(연결-분기) 날짜 데이터 엑셀에 넣기
                    e = d + 66
                    g = chr(e) + str(50)
                    wb_data[g] = ifrs_DQ_date[d + 2].text

                line = 0
                row = 69
                for b in range(0, 8):                                                 # ifrs(개별-분기) 숫자 데이터 엑셀에 넣기
                    samsung = 66
                    for c in range(0, 8):
                        rrow = str(row)
                        col = chr(samsung)
                        cell = col + rrow
                        wb_data[cell] = ifrs_BQ[line].text
                        line += 1
                        samsung += 1
                    row += 1

                for d in range(0, 5):                                                 # ifrs(개별-분기) 날짜 데이터 엑셀에 넣기
                    e = d + 66
                    g = chr(e) + str(68)
                    wb_data[g] = ifrs_BQ_date[d + 2].text


                wb.save('G:\Hyuk_Rim_v5.xlsx')                                 # 엑셀 파일 이름 바뀔때마다 업데이트!!
                wb.close()
                # 엑셀 계산시트 종료

                wb3 = xw.Book('G:\Hyuk_Rim_v5.xlsx')                          # 엑셀 파일 이름 바뀔때마다 업데이트!!
                # xlwings 이용해서 엑셀의 셀값 가져오기
                wb3_result = wb3.sheets['Result']
                wb3_data = wb3.sheets['Data']

                if wb3_result.range('C26').value >= wb3_result.range('C23').value:
                    # 매수가격 >= 현재가격, 인데 일단 테스트 용으로 뒤집어 놓음!!!
                    if wb3_result.range('I31').value >= 0.01:
                        # 배당수익률 1% 이상
                        buy_zoo.append(wb3_data.range('B4').value)
                app = xw.apps.active
                app.quit()
                print(a)
    except AttributeError:
        print('에러 = ', a)
    except IndexError:
        print('index 에러 = ', a)


wb2 = Workbook()
# 결과 기록할 엑셀 파일 만들기
ws = wb2.active
ws.title = "result"
# 엑셀 결과시트 이름 지정
yo = 2
for z in range(0, len(buy_zoo)):
    # b열에 매수할 종목들 하나씩 기록
    ws['B' + str(yo)] = buy_zoo[z]
    yo += 1

wb2.save('G:\RESULT.xlsx')
hms(time.time() - start)
# 엑셀 파일 저장, 경로까지 적으면 원하는 위치 저장 가능. 디폴트 경로는 파이썬 코드가 있는 곳
# ws.values 셀의 수식이 아닌 값만을 가져온다
# G:\Hyuk_Rim.xlsx
